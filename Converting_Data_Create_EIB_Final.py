import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re
import numpy as np
import multiprocessing
import pandas as pd
import numpy as np
import os

def get_mapped_cost_center_data(input_data, mapping_file,
                                all_unique_data_list, unavailable_reference_id_list):
    convert_and_format_date(input_data, 'Effective Date')
    input_data['Add_Only'] = 'Y'
    input_data = generate_row_id(input_data, 'System ID', 'Organization Code', 'Row ID*')
    input_data['spreadsheet_key'] = (input_data['Cost Center ID'] != input_data['Cost Center ID'].shift()).cumsum()
    print("aPNA INPUT")
    print(input_data)
    columns_mapping = {
        'spreadsheet_key': 2,
        'Add_Only': 3,
        'Effective Date': [5,14],
        #'Cost Center Name': 9,
        'Cost Center ID': [7],
        #'Organization Type Name*': 8,
        #'Organization Subtype Name*': 9,
        'System ID' : 18,
        'Organization Code': 11,
        #'Cost Center Hierarchy': 21
        'Row ID*' : 17,

    }
    # data_transform_mapping = {"Job Family ID": "Job_Family_ID"}
    # input_data, all_unique_values_list, un_available_reference_type_id = mapping_data(input_data,
    #                                                                                   data_transform_mapping,
    #                                                                                   mapping_file,
    #                                                                                   all_unique_data_list,
    #                                                                                   unavailable_reference_id_list,
    #                                                                                   "Cost Center")
    return input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list


def convert_and_format_date(input_data, column):
    # Convert column to datetime format
    input_data[column] = pd.to_datetime(input_data[column], errors='coerce')
    # Format column as "YYYY-MM-DD"
    # input_data[column] = input_data[column].dt.strftime('%Y-%b-%d')
    input_data[column] = input_data[column].dt.strftime('%Y-%m-%d')


def get_mapped_location_data(input_data, mapping_file,
                             all_unique_data_list, unavailable_reference_id_list):
    convert_and_format_date(input_data, 'Effective Date')
    input_data['Add_Only'] = 'Y'
    input_data['key'] = (input_data['Location ID'] != input_data['Location ID'].shift()).cumsum()

    columns_mapping = {
        'key': 2,
        'Add_Only': 3,
        'Effective Date': 50,
        'Location Name*': 8,
        'Location ID': 6,
        'Location Usage*+*': 9,
        'Latitude':15,
        'Longitude':16,
        'Default Currency':22,
        # 'Location Type+': 10,                     ///////////changed by darshan
        # 'Time Profile': 18,
        # 'Display Language': 20,
        # 'Time Zone': 21,
        # 'Default Currency':22,
        # 'Municipality': 57,
        # 'Country Region': 64,
        # 'Country': [51, 59],
        # 'Postal Code': 70
        #'Location Hierarchy': 154
    }

    data_transform_mapping = {
                            "Location ID": "Location_ID",
                            #   "Time Profile": "Time_Profile_ID",                  ///////////changed by darshan
                            #   "Display Language": "User_Language_ID",
                            #   "Time Zone": "Time_Zone_ID",
                            #   "Country": "ISO_3166-1_Alpha-3_Code",
                              "Location Usage*+*": "Location_Usage_ID",
                            #   "Country Region": "Country_Region_ID",
                            #   'Country':"ISO_3166-1_Alpha-3_Code"
                              #"Location Hierarchy": "Location_Hierarchy_ID"
                              }
    input_data, all_unique_values_list, un_available_reference_type_id = mapping_data(input_data,
                                                                                      data_transform_mapping,
                                                                                      mapping_file,
                                                                                      all_unique_data_list,
                                                                                      unavailable_reference_id_list,
                                                                                      "Location")
    #print("idhar hai")
    #print(input_data.head())
    return input_data, columns_mapping, all_unique_values_list, un_available_reference_type_id

def get_mapped_job_profile_data(input_data, mapping_file,
                                all_unique_data_list, unavailable_reference_id_list):
    convert_and_format_date(input_data, 'Effective Date')
    input_data['Add_Only'] = 'Y'

    input_data = generate_row_id(input_data, 'Job Code','Job Title', "Row Id*")
    columns_mapping = {
        'Spreadsheet Key*': 2,
        'Job Profile Reference':3,
        # 'Add_Only': 3,
        #'Job Profile ID':4,
        'Job Code': 4,
        'Effective Date': 5,
        'Row Id*':6,
        'Inactive':7,
        'Job Title': 8,
        'Include Job Code in Name':9,
        'Job Profile Summary': 11,
        #'Job Description': 12,
        # 'Work Shift Required': 14,
        # 'Is Job Public': 15,
        # 'Inactive': 7,
        # 'Management Level': 16,
        # 'Job Category': 17,
        # 'Job Level': 18,
        # 'Job Family*': 21,
        # 'Company Insider Type*': 24,
        # 'Referral Payment Plan': 25,
        # 'Critical Job': 26,
        # 'Difficulty to Fill': 27,
        # 'Job Classification*': 31,
        # 'Pay Rate Type Country': 34,
        # 'Pay Rate Type': 35,
        # 'Row ID': 36,
        # 'Job Profile Exempt - Country/Country Region': 38,
        # 'Job Exempt': 40,
        # 'Compensation Grade': 96,
        # 'Compensation Grade Profile': 97,
        # 'Allowed Unions': 98,
    }

    data_transform_mapping = {
                            #   "Inactive": "Numeric_Boolean_ID", "Work Shift Required": "Numeric_Boolean_ID",


                            #   "Is Job Public": "Numeric_Boolean_ID", "Management Level": "Management_Level_ID",
                            #   "Job Category": "Job_Category_ID", "Job Family*": "Job_Family_ID",
                            #   "Job Classification*": "Job_Classification_Reference_ID",
                            #   "Pay Rate Type Country": "ISO_3166-1_Alpha-3_Code", "Pay Rate Type": "Pay_Rate_Type_ID",
                            #   "Job Exempt": "Numeric_Boolean_ID",
                            #   'Job Profile Exempt - Country/Country Region':"ISO_3166-1_Alpha-3_Code"
                              }
    # input_data, all_unique_values_list, un_available_reference_type_id = mapping_data(input_data,
    #                                                                                   data_transform_mapping,
    #                                                                                   mapping_file,
    #                                                                                   all_unique_data_list,
    #                                                                                   unavailable_reference_id_list,
    #                                                                                   "Job Profile")
    # return input_data, columns_mapping, all_unique_values_list, un_available_reference_type_id
    #print("input hai jiii")
    #print(input_data.head())

    return input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list






multiprocessing.set_start_method("spawn", force=True)



def split_list(a_list):
    half = len(a_list)/2
    return a_list[:half], a_list[half:]


def convert_data(input_data, sheet, load, mapping_file, mapping_data_dict,
                 all_unique_data_list, unavailable_reference_id_list, input_file, eib_file_name):
    start_row = 6
    columns_mapping = {}

    # CountryISO2Code = "AU"
    # CountryISO3Code = "AUS"
    # date_today = datetime.now().date().strftime('%Y-%m-%d')
    currency_code = "AUD"

    mapping_file = mapping_file
    defalut_compensation_count = 0
    if load == "Change Personal Information":

        input_data = input_data.dropna(subset=['Gender', 'Date of Birth', 'Marital Status', 'Marital Status Date', 'Race/Ethnicity', 'Citizenship Status', 'Citizenship Status Country', 'Pronoun',	'Gender Identity', 'Disability Status', 'Veteran Status Name'], how='all')
        input_data.reset_index(drop=True, inplace=True)

        input_data['spreadsheet_key'] = (input_data['Legacy Worker ID'] != input_data['Legacy Worker ID'].shift()).cumsum()

        # format columns to date
        convert_and_format_date(input_data, 'Date of Birth')
        convert_and_format_date(input_data, 'Marital Status Date')

        # Blank out 'Marital Status Date' column where 'Marital Status' is NaN or blank
        input_data.loc[(input_data['Marital Status'] == 'Engaged (Australia)'), 'Marital Status Date'] = ""

        columns_mapping = {
            'spreadsheet_key': 2,
            'Legacy Worker ID': 3,
            'Date of Birth': 6,
            'Gender': 9,
            'Disability Status': 14,
            'Marital Status': 36,
            'Citizenship Status': 38,
            'Citizenship Status Country': 39,
            'Race/Ethnicity': 41,
            'Marital Status Date': 72,
            'Pronoun': 82,
            'Gender Identity': 81
        }

        data_transform_mapping = {
            'Gender': 'Gender_Code',
            'Marital Status': 'Marital_Status_ID',
            'Citizenship Status': 'Citizenship_Status_Code',
            'Race/Ethnicity': 'Ethnicity_ID',
            'Disability Status': 'Disability_ID',
            'Legacy Worker ID': "Worker_WID",
            'Gender Identity': 'Gender_Identity_ID',
            'Pronoun': 'Pronoun_ID',
            'Citizenship Status Country': 'ISO_3166-1_Alpha-3_Code'
        }
        # 'Race/Ethnicity':'Ethnicity_ID','Disability Status':'Disability_ID',

        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)
        print(">>>>", len(input_data))

    elif load == 'Cost Center':
        # from get_mapping_for_loads import get_mapped_cost_center_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_cost_center_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Location":
        # from get_mapping_for_loads import get_mapped_location_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_location_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Job Profile":
        # from get_mapping_for_loads import get_mapped_job_profile_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_profile_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        print("Heree")
        print(input_data.head(1))


    row_count = 1
    position_start_row = 6
    last_worker = 0
    last_index = 0
    emergency_priority = 1
    job_req_validation = {}
    temp_dict = {}
    comp_grad_count, comp_profile_count, comp_step_count = 1, 1, 1

    if sheet.title == 'Request Default Compensation':
        last_index = defalut_compensation_count

    # input_data = input_data.head(100)
    #input_data = input_data[input_data["Legacy Worker ID"]=="5f2619b812681001715be7c1bebc0000"]


    for idx, row in input_data.iterrows():
        if row.get('Legacy Worker ID'):
            if ((load != "Personal Work Contact Change") or (
                    load == "Personal Work Contact Change" and pd.notna(row['Work Phone Number']))):# or (
                    # load == "Work Contact Change" and pd.notna(row['Work Email']))):

                if last_worker == row['Legacy Worker ID']:
                    last_index = last_index
                    emergency_priority += 1
                else:
                    last_index += 1
                    emergency_priority =1


        else:
            last_index += 1
            emergency_priority = 1
            for column, cell_value in columns_mapping.items():
                if column in ['key', 'Spreadsheet Key*']:
                    if sheet.title != 'Edit Assign Organization':
                      if column in input_data.columns:
                        sheet.cell(row=start_row, column=cell_value, value=row[column])
                      else:
                        sheet.cell(row=start_row, column=cell_value, value=last_index)

                elif column == 'RowID':
                    if isinstance(cell_value, list):
                        cell_value_list = cell_value
                        for i, value in enumerate(cell_value_list):
                            sheet.cell(row=start_row, column=value, value=1)
                    else:
                        sheet.cell(row=start_row, column=cell_value, value=1)

                else:
                    if isinstance(cell_value, list):
                        cell_value_list = cell_value
                        for i, value in enumerate(cell_value_list):
                            sheet.cell(row=start_row, column=value, value=row[column])
                    else:
                        value = None if pd.isna(row[column]) else row[column]
                        #print(start_row, cell_value, row[column])
                        sheet.cell(row=start_row, column=cell_value, value=value)
        #print("sheet is")    
        #for r in sheet.iter_rows(values_only=True): print(r)
        if load == 'One Time Payments':
            idx += 1
            sheet.cell(row=start_row, column=8, value=idx)

        if load == 'Cost Center':
            #sheet.cell(row=start_row, column=3, value="Y")
            sheet.cell(row=start_row, column=12, value="Y")
            sheet.cell(row=start_row, column=13, value="Y")
            sheet.cell(row=start_row, column=15, value="9c875610c4fc496499e39741b6541dbc")
            sheet.cell(row=start_row, column=20, value="Org_SubType_Cost_Center")


        if load == "Location":
            idx += 1
            address_row = 1
            sheet.cell(row=start_row, column=3, value="Y")
            sheet.cell(row=start_row, column=44, value=1)
            sheet.cell(row=start_row, column=71, value=idx)
            sheet.cell(row=start_row, column=73, value=idx)
            sheet.cell(row=start_row, column=72, value="Y")
            sheet.cell(row=start_row, column=74, value="Y")
            sheet.cell(row=start_row, column=75, value="BUSINESS")
            sheet.cell(row=start_row, column=77, value="COMMUNICATION_USAGE_BEHAVIOR_TENANTED-6-10")

            # if pd.notna(row['Primary Address - Line 1']):                 ///////////changed by darshan
            #     sheet.cell(row=start_row, column=2, value=last_index)
            #     sheet.cell(row=start_row, column=53, value=address_row)
            #     sheet.cell(row=start_row, column=55, value="ADDRESS_LINE_1")
            #     sheet.cell(row=start_row, column=56, value=row['Primary Address - Line 1'])
            #     address_row += 1

            # if pd.notna(row['Primary Address - Line 2']):
            #     start_row += 1
            #     sheet.cell(row=start_row, column=2, value=last_index)
            #     sheet.cell(row=start_row, column=53, value=address_row)
            #     sheet.cell(row=start_row, column=44, value=1)
            #     sheet.cell(row=start_row, column=55, value="ADDRESS_LINE_2")
            #     sheet.cell(row=start_row, column=56, value=row['Primary Address - Line 2'])
            #     address_row += 1

            # if pd.notna(row['Primary Address - Line 3']):
            #     start_row += 1
            #     sheet.cell(row=start_row, column=2, value=last_index)
            #     sheet.cell(row=start_row, column=53, value=address_row)
            #     sheet.cell(row=start_row, column=44, value=1)
            #     sheet.cell(row=start_row, column=55, value="ADDRESS_LINE_3")
            #     sheet.cell(row=start_row, column=56, value=row['Primary Address - Line 3'])
            #     address_row += 1

            # if pd.notna(row['Primary Address - Line 4']):
            #     start_row += 1
            #     sheet.cell(row=start_row, column=2, value=last_index)
            #     sheet.cell(row=start_row, column=53, value=address_row)
            #     sheet.cell(row=start_row, column=44, value=1)
            #     sheet.cell(row=start_row, column=55, value="ADDRESS_LINE_4")
            #     sheet.cell(row=start_row, column=56, value=row['Primary Address - Line 4'])
            #     address_row += 1

        if load == "Job Profile":
            idx += 1

            sheet.cell(row=start_row, column=7, value="n")

            # if pd.notna(row['Job Family*']):
            #     sheet.cell(row=start_row, column=19, value=idx)
            #     sheet.cell(row=start_row, column=21, value=row['Job Family*'])

            # if pd.notna(row['Job Classification*']):
            #     sheet.cell(row=start_row, column=29, value=idx)
            #     sheet.cell(row=start_row, column=31, value=row['Job Classification*'])

            # if pd.notna(row['Pay Rate Type']):
            #     sheet.cell(row=start_row, column=32, value=idx)
            #     sheet.cell(row=start_row, column=32, value=row['Pay Rate Type Country'])
            #     sheet.cell(row=start_row, column=35, value=row['Pay Rate Type'])

            # if pd.notna(row['Job Exempt']) and row['Job Exempt']=="Y":
            #     sheet.cell(row=start_row, column=36, value=idx)
            #     sheet.cell(row=start_row, column=38, value=row['Job Profile Exempt - Country/Country Region'])
            #     sheet.cell(row=start_row, column=40, value=row['Job Exempt'])

            # if pd.notna(row["Workers' Compensation Code"]):
            #     sheet.cell(row=start_row, column=41, value=idx)
            #     sheet.cell(row=start_row, column=42, value=row["Workers' Compensation Code"])

            # if pd.notna(row["Responsibilities"]):
            #     sheet.cell(row=start_row, column=44, value=idx)
            #     sheet.cell(row=start_row, column=45, value=row["Responsibilities"])

            # if pd.notna(row["Work Experience"]):
            #     sheet.cell(row=start_row, column=48, value=idx)
            #     sheet.cell(row=start_row, column=49, value=row["Work Experience"])

            # if pd.notna(row["Education"]):
            #     sheet.cell(row=start_row, column=53, value=idx)
            #     sheet.cell(row=start_row, column=54, value=row["Education"])

            # if pd.notna(row["Languages"]):
            #     sheet.cell(row=start_row, column=58, value=idx)
            #     sheet.cell(row=start_row, column=59, value=row["Languages"])

            # if pd.notna(row["Competencies"]):
            #     sheet.cell(row=start_row, column=65, value=idx)
            #     sheet.cell(row=start_row, column=66, value=row["Competencies"])

            # if pd.notna(row["Certifications"]):
            #     sheet.cell(row=start_row, column=70, value=idx)
            #     sheet.cell(row=start_row, column=72, value=row["Certifications"])

            # if pd.notna(row["Training"]):
            #     sheet.cell(row=start_row, column=80, value=idx)
            #     sheet.cell(row=start_row, column=81, value=row["Training"])

            # if pd.notna(row["Skills on Job Profile"]):
            #     sheet.cell(row=start_row, column=86, value=idx)
            #     sheet.cell(row=start_row, column=87, value=row["Skills on Job Profile"])

            # if pd.notna(row["Integration Identifier"]):
            #     sheet.cell(row=start_row, column=92, value=idx)
            #     sheet.cell(row=start_row, column=93, value=row["Integration Identifier"])

        start_row += 1

    return sheet, all_unique_data_list, unavailable_reference_id_list

def generate_row_id(df, groupby_col, sort_col,row_id_col):
    """
    Generate 'row id' that increments only when the 'sort_col' value changes between
    the current row and the previous row within each group of 'groupby_col'.

    Args:
    - df: DataFrame containing the data
    - groupby_col: Column to group by ('Legacy Worker ID' in this case)
    - sort_col: Column to use for sorting ('test' in this case)

    Returns:
    - DataFrame with 'row id' column added
    """
    df[row_id_col] = (df.groupby(groupby_col)[sort_col]
                    .apply(lambda x: (x != x.shift()).cumsum())
                    .reset_index(drop=True))

    # Blank out the 'row_id_col' if df[sort_col] is NaN or blank
    df.loc[df[sort_col].isnull() | (df[sort_col] == ''), row_id_col] = None

    return df



def load_file_params(load, base_path):

    # Define the dictionary with load conditions as keys and corresponding values
    params = {
        "Change Personal Information": {
            "input_sheet": "Personal-Info",
            "eib_file_name": ""
        },"Location": {
            "input_sheet": "Location",
            "eib_file_name": base_path + "EIB/Put_Location_v46.0.xlsx"
        },
        "Job Profile": {
            "input_sheet": "Job-Profile",
            "eib_file_name": base_path + "EIB/Submit_Job_Profile_v46.0.xlsx"
        },
        "Cost Center": {
            "input_sheet": "Cost-Center",
            "eib_file_name": base_path + "EIB/Put_Cost_Center_v46.0.xlsx"
        },
    }

    params_for_load = params.get(load, {})

    # Extract the input_sheet and eib_file_name values
    input_sheet = params_for_load.get("input_sheet", "")
    eib_file_name = params_for_load.get("eib_file_name", "")

    # Return the values
    return input_sheet, eib_file_name



def get_mapping_data_dict(mapping_file):
    mapped_data = pd.read_excel(mapping_file)
    mapping_data_dict = {}
    for idx, row in mapped_data.iterrows():
        value_dict = {row['Reference ID Type']: row['Reference ID Value']}
        mapping_data_dict[row['Business Object Instance']] = value_dict

    return mapping_data_dict


from openpyxl import load_workbook


def validate_eib_sections(eib_file_path):

    """print("======================================")
    print("Opening Workbook:", eib_file_path)
    print("======================================")"""

    wb = load_workbook(eib_file_path)
    sheet = wb.active

    AREA_ROW = 2
    REQ_ROW = 3
    FIELD_ROW = 5
    DATA_START_ROW = 6

    """print("\nSheet Name:", sheet.title)
    print("Total Columns:", sheet.max_column)
    print("Total Rows:", sheet.max_row)"""

    sections = {}
    current_area = None

   # print("\n======================================")
    #print("STEP 1: Detect Sections")
    #print("======================================")

    for col in range(1, sheet.max_column + 1):

        area_val = sheet.cell(row=AREA_ROW, column=col).value
        restriction_val = sheet.cell(row=REQ_ROW, column=col).value
        field_name = sheet.cell(row=FIELD_ROW, column=col).value

       # print("\n--- COLUMN", col, "---")
       # print("Raw Area Value:", area_val)
        #print("Format Value:", format_val)
        #print("Field Name:", field_name)

        # forward fill merged area cells
        if area_val:
            current_area = area_val

        area = current_area

        #print("Resolved Area:", area)

        if not field_name:
            #print("Skipping column because field_name is empty")
            continue

        if area not in sections:
            sections[area] = []

        is_required = False
        if restriction_val and "required" in str(restriction_val).lower():
            is_required = True

        #print("Is Required:", is_required)

        sections[area].append({
            "col": col,
            "field": field_name,
            "required": is_required
        })

    #print("\n======================================")
  #  print("SECTIONS DETECTED")
   # print("======================================")

    #for sec, cols in sections.items():
      #  print("\nSECTION:", sec)
        #for c in cols:
          #  print("   Column:", c["col"], "| Field:", c["field"], "| Required:", c["required"])

    errors = []
    warnings = []

    #print("\n======================================")
    #print("STEP 2: Validate Rows")
    #print("======================================")

    for row in range(DATA_START_ROW, sheet.max_row + 1):

        #print("\n--------------------------------------")
       # print("Checking Row:", row)
        #print("--------------------------------------")

        for section, columns in sections.items():

            #print("\nSECTION:", section)

            section_values = []
            required_values = []
            required_fields = []

            for col_info in columns:

                value = sheet.cell(row=row, column=col_info["col"]).value

                """print(
                    "Column:",
                    col_info["col"],
                    "| Field:",
                    col_info["field"],
                    "| Value:",
                    value,
                    "| Required:",
                    col_info["required"]
                )l"""

                section_values.append(value)

                if col_info["required"]:
                    required_values.append(value)
                    required_fields.append(col_info["field"])

            #print("Section Values:", section_values)
           # print("Required Values:", required_values)

            # Section completely empty
            if all(v is None or str(v).strip() == "" for v in section_values):
                #print("Section completely empty → SKIP")
                continue

            #print("Section has data → validating required fields")

            filled_required = [
                v for v in required_values if v not in (None, "", " ")
            ]

            #print("Filled Required Count:", len(filled_required))
            #print("Total Required Fields:", len(required_values))

            # ERROR: section has data but all required fields empty
            if len(filled_required) == 0 and len(required_values) > 0:

                #print("ERROR detected in section:", section)

                errors.append({
                    "row": row,
                    "section": section,
                    "missing_fields": required_fields
                })

            # WARNING: partially filled required fields
            elif 0 < len(filled_required) < len(required_values):

                #print("WARNING detected in section:", section)

                warnings.append({
                    "row": row,
                    "section": section,
                    "missing_fields": [
                        f for f, v in zip(required_fields, required_values)
                        if v in (None, "", " ")
                    ]
                })


    """print("======================================")
    print("FINAL RESULT")
    print("======================================")

    print("Errors:", errors)
    print("Warnings:", warnings)"""

    return errors, warnings
   
   




def process_load(load, all_unique_data_list, unavailable_reference_id_list, mapping_file,input_data_temp_file,input_sheet=None):

    #print("First row of input file:")
    #print(input_data_temp_file.iloc[1])
    # mapping_file = base_path + 'Combined_Mapping.xlsx'
    eib_file_name = ''
    if not input_sheet:
        input_sheet, eib_file_name = load_file_params(load, base_path)

    # read Excel sheets
    #if load == "Absence Input":
        #input_data = pd.read_excel(input_data_temp_file, sheet_name=input_sheet, skiprows=[0, 1, 3, 4], dtype=object)
    #elif load == 'Worker Additional Data':
        #input_data = pd.read_excel(input_data_temp_file, sheet_name=input_sheet, skiprows=[0, 2, 3], dtype=object)
    #else:


    # input_data = pd.read_excel(input_data_temp_file, sheet_name=input_sheet, skiprows=[1, 2], dtype=object) ///////////changed by darshan
    print("Input sheet is : ", input_sheet)
    input_data = pd.read_excel(input_data_temp_file, sheet_name=input_sheet, dtype=object)
    #input_data.head()
    print("Input data is", input_data.iloc[1])

    # initiate and laod the eib template that we need to write
    wb = load_workbook(eib_file_name)
    sheet_names = wb.sheetnames
    sheet_list = []
    if load == "Change Job Assign Pay Group":
        sheet_name = wb[sheet_names[6]]
        sheet_list.append(sheet_name)

    elif load in ("Payment Elections", "Absence Input", "Supervisory Org", "Prehire", "Put Candidate", "Supplier",
                  "Add Workday Account", "Company", "Update Workday Account","Job Classification", "Job Family",
                  "Job Family Group", "Cost Center", "Location Hierarchy", "Custom Organizations", "Cost Center Hierarchy", "Collective Agreement",
                   "Put Supervisory Assignment Restrictions", "Location", "Job Category", "Comp Grade and Grade Profile", "Skills",
                   "Job History Company", "Skills Reference Data", "Future Prehire"):
        sheet_name = wb[sheet_names[0]]
        sheet_list.append(sheet_name)


    else:
        sheet_name = wb[sheet_names[1]]
        sheet_list.append(sheet_name)

    mapping_data_dict = get_mapping_data_dict(mapping_file)

    for sheet in sheet_list:
        # delete all the data from sheet from 6th row
        sheet.delete_rows(6, sheet.max_row)

        # method to convert our data and write it back to sheet
        sheet, all_unique_data_list, unavailable_reference_id_list = convert_data(input_data, sheet, load, mapping_file,
                                                                                  mapping_data_dict,
                                                                                  all_unique_data_list,
                                                                                  unavailable_reference_id_list,
                                                                                  input_data_temp_file, eib_file_name)
                                                                    
        #print("dekhte hai bhai")
       # print(type(sheet.cell(row=start_row, column=cell_value).value))
        #print(sheet.cell(row=start_row, column=cell_value).value)

        # save the eib file and mark the load complete
        wb.save(eib_file_name)
        print("Completed " + load + " load")

    return all_unique_data_list, unavailable_reference_id_list





def creating_eib_files_v1(all_loads, local_base_path, mapping_file, input_data_temp_file, return_all_unique_data_list, return_unavailable_reference_id_list):

    global base_path
    base_path = local_base_path
    all_unique_data_list, unavailable_reference_id_list = pd.DataFrame(), pd.DataFrame()

    # hire, job req, request comp change, contingent worker hire, change job for contingent
    for load in all_loads:
        print("Starting load for: " + load)
        all_unique_data_list, unavailable_reference_id_list = process_load(load, all_unique_data_list,
                                                                           unavailable_reference_id_list, mapping_file,input_data_temp_file)

    # all_unique_data_list.dropna(subset=['Unique Value','Mapped Value'], inplace=True)
    if 'Unique Value' in all_unique_data_list.columns and 'Mapped Value' in all_unique_data_list.columns:
        all_unique_data_list.dropna(subset=['Unique Value','Mapped Value'], how='all', inplace=True)

    count_null_values = 0
    if 'Mapped Value' in all_unique_data_list.keys():
        count_null_values = all_unique_data_list['Mapped Value'].isnull().sum()

    if count_null_values > 0:
        print("Error: There are {} records where the value mapping is blank or null.".format(count_null_values))

    record_count = len(unavailable_reference_id_list)

    if record_count > 0:
        print("Error: The are {} mapping could not be found in the mapping file.".format(record_count))

    return_all_unique_data_list.append(all_unique_data_list)
    return_unavailable_reference_id_list.append(unavailable_reference_id_list)

    return return_all_unique_data_list, return_unavailable_reference_id_list


def mapping_data(input_data, data_transform_mapping, mapping_file, all_unique_data_list, unavailable_reference_id_list,
                 load):
    # global all_unique_data_list  # Add global keyword
    mapping_data = pd.read_excel(mapping_file)

    # Map the values based on the mapping file and handle KeyError
    unique_values = pd.DataFrame(
        columns=['Unique Value', 'Mapped Value', 'Mapped Reference ID Type', 'Load'])

    un_mapped_values = pd.DataFrame(
        columns=['Unique Value', 'UnMapped Reference ID Type', 'Load'])

    for column, filter_value in data_transform_mapping.items():
        # input_data[column] = input_data[column].astype('str')
        # convert the column to str if it not already
        mapped_data = \
            mapping_data[mapping_data['Reference ID Type'] == filter_value].set_index('Business Object Instance')[
                'Reference ID Value']

        un_available_id_type = pd.DataFrame()
        if not len(mapped_data):
            un_mapped_values = pd.concat([
                un_mapped_values,
                pd.DataFrame({
                    'Unique Value': [column],
                    'UnMapped Reference ID Type': [filter_value],
                    'Load': [load]
                })
            ], ignore_index=True)

        input_data = pd.merge(input_data, mapped_data, left_on=column, right_on='Business Object Instance', how='left',
                              suffixes=('', '_mapped'))

        # create new dataframe with value and mapped value and collate together to form giant document
        unique_column_values = input_data[[column, 'Reference ID Value']].drop_duplicates()
        unmapped_id = 0
        # if not len(mapped_data):
        #     unmapped_id = filter_value
        unique_values = pd.concat([
            unique_values,
            pd.DataFrame({
                'Unique Value': unique_column_values[column],
                'Mapped Value': unique_column_values['Reference ID Value'],
                'Mapped Reference ID Type': filter_value,
                'Load': load
            })
        ], ignore_index=True)

        # replace the column value with mapped value in input_data dataframe

        input_data[column] = input_data['Reference ID Value']
        input_data.drop('Reference ID Value', axis=1, inplace=True)
        # input_data[column] = input_data[column].map(mapped_data)

    # # Append the unique values to the global list
    all_unique_data_list = pd.concat([all_unique_data_list, unique_values], ignore_index=True)
    unavailable_reference_id_list = pd.concat([unavailable_reference_id_list, un_mapped_values],
                                              ignore_index=True)
    # input_data.drop_duplicates(keep=False, inplace=True)
    return input_data, all_unique_data_list, unavailable_reference_id_list

def creating_eib_files_with_parallel_processing(all_loads, local_base_path, mapping_file, input_data_temp_file):
    manager = multiprocessing.Manager()

    return_all_unique_data_list = manager.list()
    return_unavailable_reference_id_list = manager.list()
    # creating processes
    if len(all_loads) > 1:

        p1_all_loads, p2_all_loads = split_list(all_loads)
        print('>>>>>', p1_all_loads, p2_all_loads)

        p1 = multiprocessing.Process(target=creating_eib_files_v1, args=(p1_all_loads, local_base_path, mapping_file, input_data_temp_file, return_all_unique_data_list, return_unavailable_reference_id_list,))
        p2 = multiprocessing.Process(target=creating_eib_files_v1, args=(p2_all_loads, local_base_path, mapping_file, input_data_temp_file, return_all_unique_data_list, return_unavailable_reference_id_list,))

        # starting process 1
        p1.start()
        # starting process 2
        p2.start()

        # wait until process 1 is finished
        p1.join()

        # wait until process 2 is finished
        p2.join()

        print("Both processes finished Done!")
    else:
        creating_eib_files_v1(all_loads, local_base_path, mapping_file, input_data_temp_file, return_all_unique_data_list, return_unavailable_reference_id_list,)

    list_df_all_unique_data_list = [n for n in return_all_unique_data_list]

    list_df_unavailable_reference_id_list = [n for n in return_unavailable_reference_id_list]

    final_df_all_unique = pd.concat(list_df_all_unique_data_list, axis=0, ignore_index=True)
    final_df_unavailable_id = pd.concat(list_df_unavailable_reference_id_list, axis=0, ignore_index=True)

    with pd.ExcelWriter(r'C:\Users\patel\Desktop\final code\uniquefile.xlsx', engine='xlsxwriter') as writer:
        final_df_all_unique.to_excel(writer, sheet_name='unique_mapped_values', index=False)
        final_df_unavailable_id.to_excel(writer, sheet_name='un_available_reference_type_id', index=False)

    print("EIB Files Creation Completed...")